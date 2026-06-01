from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.conf import settings

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.authtoken.models import Token

from .models import Project, TaskHistory
from .serializers import RegisterSerializer, ProjectSerializer, TaskHistorySerializer

import pandas as pd
import json
import os
import tempfile


# ── 회원가입 ──────────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        token, _ = Token.objects.get_or_create(user=user)
        return Response({
            'token': token.key,
            'username': user.username,
        }, status=201)
    return Response(serializer.errors, status=400)


# ── 로그인 ────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    username = request.data.get('username')
    password = request.data.get('password')

    user = authenticate(request, username=username, password=password)
    if not user:
        return Response({'error': '아이디 또는 비밀번호가 틀렸습니다.'}, status=401)

    token, _ = Token.objects.get_or_create(user=user)
    return Response({
        'token': token.key,
        'username': user.username,
    })


# ── 로그아웃 ──────────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    request.user.auth_token.delete()
    return Response({'message': '로그아웃 됐어요.'})


# ── 내 정보 확인 ──────────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me(request):
    return Response({
        'username': request.user.username,
        'user_id': request.user.id,
    })


# ── Excel 처리 유틸 함수 ──────────────────────────────
def col_letter_to_index(col):
    col = col.strip().upper()
    result = 0
    for ch in col:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1


def process_excel(input_file, output_file, input_sheet, output_sheet,
                  matching_cols, output_cols, transfer_pairs, filter_settings):
    df_input = pd.read_excel(input_file, sheet_name=input_sheet or 0, dtype=str).fillna('')

    for col, values in filter_settings.items():
        if values:
            idx = col_letter_to_index(col)
            if idx < df_input.shape[1]:
                df_input = df_input[df_input.iloc[:, idx].isin(values)]

    output_bytes = output_file.read()
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(output_bytes)
        tmp_path = tmp.name

    df_output = pd.read_excel(tmp_path, sheet_name=output_sheet or 0, dtype=str).fillna('')

    matching_cols = [c.strip() for c in matching_cols if c.strip()]
    output_cols   = [c.strip() for c in output_cols if c.strip()]

    for _, input_row in df_input.iterrows():
        for out_idx, out_row in df_output.iterrows():
            match = all(
                input_row.iloc[col_letter_to_index(mc)] == out_row.iloc[col_letter_to_index(oc)]
                for mc, oc in zip(matching_cols, output_cols)
                if col_letter_to_index(mc) < df_input.shape[1]
                and col_letter_to_index(oc) < df_output.shape[1]
            )
            if match:
                for from_col, to_col in transfer_pairs:
                    fi = col_letter_to_index(from_col)
                    ti = col_letter_to_index(to_col)
                    if fi < df_input.shape[1] and ti < df_output.shape[1]:
                        df_output.iat[out_idx, ti] = input_row.iloc[fi]

    result_filename = f'outputs/result_{os.path.basename(tmp_path)}'
    result_path = os.path.join(settings.MEDIA_ROOT, result_filename)
    os.makedirs(os.path.dirname(result_path), exist_ok=True)
    df_output.to_excel(result_path, index=False)
    os.unlink(tmp_path)

    return result_filename

# ── 프로젝트 목록 조회 + 생성 ─────────────────────────
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def project_list(request):
    if request.method == 'GET':
        projects = Project.objects.filter(user=request.user).order_by('-created_at')
        serializer = ProjectSerializer(projects, many=True)
        return Response(serializer.data)

    if request.method == 'POST':
        serializer = ProjectSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)


# ── 프로젝트 삭제 ─────────────────────────────────────
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def project_delete(request, pk):
    try:
        project = Project.objects.get(pk=pk, user=request.user)
    except Project.DoesNotExist:
        return Response({'error': '프로젝트를 찾을 수 없어요.'}, status=404)
    project.delete()
    return Response({'message': '삭제됐어요.'})


# ── 시트 목록 조회 ─────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_sheets(request):
    file = request.FILES.get('file')
    if not file:
        return Response({'error': '파일이 없어요.'}, status=400)

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        for chunk in file.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    try:
        xl = pd.ExcelFile(tmp_path)
        return Response({'sheets': xl.sheet_names})
    except Exception as e:
        return Response({'error': str(e)}, status=400)
    finally:
        os.unlink(tmp_path)


# ── 파일 미리보기 (상위 20행) ──────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_preview(request):
    file  = request.FILES.get('file')
    sheet = request.data.get('sheet', '')

    if not file:
        return Response({'error': '파일이 없어요.'}, status=400)

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        for chunk in file.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    try:
        df = pd.read_excel(tmp_path, sheet_name=sheet or 0, dtype=str).fillna('')
        preview = df.head(20)

        # 컬럼 레터 생성 (A, B, C ...)
        col_letters = [chr(65 + i) for i in range(len(preview.columns))]
        col_names   = [str(c) if 'unnamed' not in str(c).lower() else '' for c in preview.columns]

        return Response({
            'col_letters': col_letters,
            'col_names':   col_names,
            'rows':        preview.values.tolist(),
        })
    except Exception as e:
        return Response({'error': str(e)}, status=400)
    finally:
        os.unlink(tmp_path)


# ── 조건 필터 값 조회 ──────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_filter_values(request):
    file  = request.FILES.get('file')
    sheet = request.data.get('sheet', '')
    cols  = [c.strip() for c in request.data.get('cols', '').split(',') if c.strip()]

    if not file:
        return Response({'error': '파일이 없어요.'}, status=400)

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        for chunk in file.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    try:
        df = pd.read_excel(tmp_path, sheet_name=sheet or 0, dtype=str).fillna('')
        result = {}
        for col in cols:
            idx = col_letter_to_index(col)
            if idx is not None and idx < df.shape[1]:
                values = sorted(set(df.iloc[:, idx].dropna().astype(str)))
                result[col] = [v for v in values if v.strip()]
        return Response(result)
    except Exception as e:
        return Response({'error': str(e)}, status=400)
    finally:
        os.unlink(tmp_path)
        


# ── Compare ────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def compare(request):
    input_file  = request.FILES.get('input_file')
    output_file = request.FILES.get('output_file')

    if not input_file or not output_file:
        return Response({'error': '파일 2개가 모두 필요해요.'}, status=400)

    try:
        input_cols      = json.loads(request.data.get('input_cols',      '[]'))
        output_cols     = json.loads(request.data.get('output_cols',     '[]'))
        filter_settings = json.loads(request.data.get('filter_settings', '{}'))
        input_sheet     = request.data.get('input_sheet',  '')
        output_sheet    = request.data.get('output_sheet', '')
    except json.JSONDecodeError as e:
        return Response({'error': f'JSON 오류: {e}'}, status=400)

    in_tmp = out_tmp = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            for chunk in input_file.chunks(): f.write(chunk)
            in_tmp = f.name

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            for chunk in output_file.chunks(): f.write(chunk)
            out_tmp = f.name

        # Input 로드 + 필터 적용
        df_in = pd.read_excel(in_tmp, sheet_name=input_sheet or 0, dtype=str).fillna('')
        for col, values in filter_settings.items():
            if values:
                idx = col_letter_to_index(col)
                if idx is not None and idx < df_in.shape[1]:
                    df_in = df_in[df_in.iloc[:, idx].isin(values)]

        # Output 로드
        df_out = pd.read_excel(out_tmp, sheet_name=output_sheet or 0, dtype=str).fillna('')

        # 복합 키 생성
        df_in['__key__']  = df_in.apply(
            lambda r: '_'.join([str(r.iloc[col_letter_to_index(c)]) for c in input_cols]), axis=1)
        df_out['__key__'] = df_out.apply(
            lambda r: '_'.join([str(r.iloc[col_letter_to_index(c)]) for c in output_cols]), axis=1)

        # 매칭
        matched = set(df_in['__key__']) & set(df_out['__key__'])
        matched_rows = df_out[df_out['__key__'].isin(matched)]

        # 중복 키 찾기
        out_dupes = {k: int(v) for k, v in
                     df_out['__key__'].value_counts().items() if v > 1 and k in matched}
        in_dupes  = {k: int(v) for k, v in
                     df_in['__key__'].value_counts().items()  if v > 1}

        return Response({
            'input_row_count':       len(df_in),
            'matched_key_count':     matched_rows['__key__'].nunique(),
            'output_duplicate_keys': out_dupes,
            'input_duplicate_keys':  in_dupes,
        })

    except Exception as e:
        return Response({'error': str(e)}, status=500)
    finally:
        for p in [in_tmp, out_tmp]:
            if p and os.path.exists(p): os.unlink(p)


# ── Save ───────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def save_result(request):
    input_file  = request.FILES.get('input_file')
    output_file = request.FILES.get('output_file')

    if not input_file or not output_file:
        return Response({'error': '파일 2개가 모두 필요해요.'}, status=400)

    try:
        input_cols      = json.loads(request.data.get('input_cols',      '[]'))
        output_cols     = json.loads(request.data.get('output_cols',     '[]'))
        transfer_pairs  = json.loads(request.data.get('transfer_pairs',  '[]'))
        custom_fills    = json.loads(request.data.get('custom_fills',    '[]'))
        filter_settings = json.loads(request.data.get('filter_settings', '{}'))
        input_sheet     = request.data.get('input_sheet',  '')
        output_sheet    = request.data.get('output_sheet', '')
    except json.JSONDecodeError as e:
        return Response({'error': f'JSON 오류: {e}'}, status=400)

    in_tmp = out_tmp = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            for chunk in input_file.chunks(): f.write(chunk)
            in_tmp = f.name

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            for chunk in output_file.chunks(): f.write(chunk)
            out_tmp = f.name

        # Input 로드 + 필터
        df_in = pd.read_excel(in_tmp, sheet_name=input_sheet or 0, dtype=str).fillna('')
        for col, values in filter_settings.items():
            if values:
                idx = col_letter_to_index(col)
                if idx is not None and idx < df_in.shape[1]:
                    df_in = df_in[df_in.iloc[:, idx].isin(values)]

        df_in['__key__'] = df_in.apply(
            lambda r: '_'.join([str(r.iloc[col_letter_to_index(c)]) for c in input_cols]), axis=1)

        # Output 로드
        import openpyxl
        from openpyxl.styles import PatternFill
        wb  = openpyxl.load_workbook(out_tmp)
        ws  = wb[output_sheet] if output_sheet and output_sheet in wb.sheetnames else wb.active
        df_out = pd.read_excel(out_tmp, sheet_name=output_sheet or 0, dtype=str).fillna('')
        df_out['__key__'] = df_out.apply(
            lambda r: '_'.join([str(r.iloc[col_letter_to_index(c)]) for c in output_cols]), axis=1)

        matched_keys  = set(df_in['__key__'])
        duplicate_keys = {k for k, v in df_in['__key__'].value_counts().items() if v > 1}
        pink = PatternFill(start_color='FFB3C6', end_color='FFB3C6', fill_type='solid')
        red  = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')

        updated = 0
        for i, key in enumerate(df_out['__key__']):
            if key not in matched_keys:
                continue
            row_num = i + 2  # 1행 헤더, 2행부터 데이터

            if key in duplicate_keys:
                for pair in transfer_pairs:
                    to_idx = col_letter_to_index(pair['to_col']) + 1
                    cell = ws.cell(row=row_num, column=to_idx)
                    cell.value = '기준 데이터 중복/ 사용자 확인필요'
                    cell.fill  = red
            else:
                input_row = df_in[df_in['__key__'] == key].iloc[0]
                for pair in transfer_pairs:
                    fi = col_letter_to_index(pair['from_col'])
                    ti = col_letter_to_index(pair['to_col']) + 1
                    if fi < df_in.shape[1]:
                        cell = ws.cell(row=row_num, column=ti)
                        cell.value = str(input_row.iloc[fi])
                        cell.fill  = pink

            for cf in custom_fills:
                ci = col_letter_to_index(cf['col']) + 1
                cell = ws.cell(row=row_num, column=ci)
                cell.value = cf['value']
                cell.fill  = pink

            updated += 1

        # 결과 파일 저장
        import uuid
        result_dir  = os.path.join(settings.MEDIA_ROOT, 'outputs')
        os.makedirs(result_dir, exist_ok=True)
        result_path = os.path.join(result_dir, f'result_{uuid.uuid4().hex[:8]}.xlsx')
        wb.save(result_path)
        wb.close()

        from django.http import FileResponse
        response = FileResponse(
            open(result_path, 'rb'),
            as_attachment=True,
            filename=os.path.basename(result_path),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['X-Updated-Count'] = updated
        response['Access-Control-Expose-Headers'] = 'X-Updated-Count'
        return response

    except Exception as e:
        return Response({'error': str(e)}, status=500)
    finally:
        for p in [in_tmp, out_tmp]:
            if p and os.path.exists(p): os.unlink(p)

@api_view(['DELETE', 'PUT'])
@permission_classes([IsAuthenticated])
def project_delete(request, pk):
    try:
        project = Project.objects.get(pk=pk, user=request.user)
    except Project.DoesNotExist:
        return Response({'error': '프로젝트를 찾을 수 없어요.'}, status=404)

    if request.method == 'DELETE':
        project.delete()
        return Response({'message': '삭제됐어요.'})

    if request.method == 'PUT':
        serializer = ProjectSerializer(project, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)        